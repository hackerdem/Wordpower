
import re
from tkinter import *
from tkinter.filedialog import askopenfile
import time
import random
import datetime
from openpyxl import load_workbook
from apscheduler.schedulers.blocking import BlockingScheduler
import pyttsx3 as vol
import os,sys
from unidecode import unidecode as udec
import threading

def ask_to_choose_import_file():
    filename = askopenfile()    
    new_data_source=filename.name
    print(new_data_source)    
    import_data_from_excel(new_data_source)

def write_imported_data_to_datatxt(source_excel_file):
    sheet_list=source_excel_file.get_sheet_names()
    word_sheet=source_excel_file.get_sheet_by_name(sheet_list[0])  
    number_of_rows=word_sheet.max_row
    print(number_of_rows)    
    with open('data.txt','a+') as newfile:#look at number of files in folder user chooses the name of file do it later
        newfile.seek(0,0)
        for i in range(1,number_of_rows):
            print(word_sheet.cell(row=i,column=1).value)
            newfile.write('::{}::{}::0::1::\n'.format(word_sheet.cell(row=i,column=1).value,word_sheet.cell(row=i,column=2).value))
        newfile.close()            
def import_data_from_excel(source):
    #let people know about formatting for excel
    try:
        source_excel_file=load_workbook(source)
        print(source)
        write_imported_data_to_datatxt(source_excel_file)        
        # add how to save file        
    except Exception as error:
        print(error)
        
def fixWord(word):
    # this section added sice pronounciation is not good for french words containing signs
    
    return udec(word)  
def fix_word(line):
    # eleminates unnecesssary double spaces and punctations
    unwanted_punctations=['  ','.','?','!',';','+','=']
    for i in unwanted_punctations:
        line= line.replace(i,'') if i in line else line
    return line 
def pronounce_meaning(mean):
    #print(volume_value.get())
    engine1=vol.init()
    engine1.setProperty('rate',180)
    engine1.setProperty('volume',volume_value.get())
    v=engine1.getProperty('voices')
    engine1.setProperty('voice',v[1].id)   
    engine1.say('means')
    time.sleep(1)    
    engine1.say(mean)
    engine1.runAndWait()
    engine1.stop()
    
def pronounce(item):
    word=item[0]
    mean=item[1]
   
    if volume_value.get()!='0.0':
        engine=vol.init()
        engine.setProperty('rate',200)
        engine.setProperty('volume',volume_value.get())
        engine.say(fixWord(word))
        engine.runAndWait() 
        engine.stop()
        time.sleep(1)
        pronounce_meaning(mean)
    else:
        time.sleep(5)
    
def get_words():
    global word_dictionary,selected_words,discard_words
    word_dictionary=[]
    selected_words=[]
    discard_words=[]
    a=0
    global size
    def append_read_data(r_file,word_list):
        for line in r_file:
            try:
                matching_pattern_in_line=re.findall('[\w, \'-]+',fix_word(line))
                print(matching_pattern_in_line)
                word_list.append(matching_pattern_in_line)
            except:
                pass 
            #implement this later
        word_list=random.shuffle(word_list)        
    try:
        print(os.path.isfile("data.txt"),os.stat("data.txt").st_size)
        if os.path.isfile("data.txt")==False or os.stat("data.txt").st_size==0:#fix this later
            import_data_from_excel(default_source_file)
    except Exception as e:
            print(e)            
    
    try:
        with open('data.txt','r') as fl:
            read_file=fl.readlines()
            size=len(read_file)
            fl.close()
    except Exception as e:
        print(e)
    append_read_data(read_file,word_dictionary)   
    with open('discarded_data.txt','r+') as discard_file:
        read_discard_file=discard_file.readlines()
        append_read_data(read_discard_file,discard_words)   
                              
    random_start=random.randrange(1,size,1)
    print(len(word_dictionary),random_start)
    for i in range(0,25):
        iterator=random_start+i if random_start+i<size else random_start+i-size
        selected_words.append(word_dictionary[iterator])
    update_list_and_interface(selected_words,word_dictionary)
    
def update_list_and_interface(selected_words,word_dictionary):
    
        for item in selected_words:
            while pause_option.get()==1:
                p=1 #just for doing nothing
            try:
                if item[2]!=None and int(item[2])>99:
                    discard_words.append(item)
                else:
                    french_word.set(item[0])
                    english_word.set(item[1])
                    item[2]='1' if item[2] in [None,0] else str(int(item[2])+1)
                    root.update()
                
                    #update_source_file(item,word_dictionary)
                    for i in word_dictionary:
                        if i[0]==item[0]:
                            i[2]=int(i[2])+int(item[2])                
                    pronounce(item)
            except Exception as error:
                discard_words.append(item)
                print(error)
                pass
            
        exit_from_application()
    
def volume_on_off():
  
    if volume_value.get()!='0.0':
        volume_value.set('0.0') 
        volume_caption.set("Volume Off") 
    else:
        volume_value.set('0.7')
        volume_caption.set("Volume On")
        
def user_discard(fr,eng):
    print(fr,eng)

    for i in word_dictionary:
        if i[0]==fr:
            discard_words.append(i)
def pause():
    #os.system("pause")
    #os.system("exec") implement resume in OS
    if pause_option.get()==0:
        pause_option.set(1)
        pause_caption.set('Resume')
    else:
        pause_option.set(0)
        pause_caption.set('Pause')
def exit_from_application():
    delete_from_source()
    os._exit(1)
    
    
def delete_from_source():
    
    try:
        with open('data.txt', 'w'): pass # empty data file before appending the list
        with open('data.txt','a+') as write_source_file:
            write_source_file.seek(0,0)# delete this later
            for i in word_dictionary:
                if i not in discard_words:
                    write_source_file.write('::{}::{}::{}::{}::\n'.format(i[0],i[1],i[2],i[3]))
                else:
                    with open('discarded_data.txt','a+') as discard_file:
                        discard_file.write('::{}::{}::{}::{}::\n'.format(i[0],i[1],i[2],i[3]))
                        
    except Exception as error:
        print(error)    
def placement(win):
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    win.geometry('{}x{}+{}+{}'.format(width,height, win.winfo_screenwidth()-width,win.winfo_screenheight()-height))          
        
def main():
    #w=threading.Thread(target=get_words).start()
    PROGRAM_NAME='FRENCH WORD LEARNING'
    global root
    global english_word
    global french_word 
    global volume_value,volume_caption    
    global pause_option,pause_caption
    global default_source_file
    root=Tk()
    default_source_file='./wordList.xlsx'
    #root.overrideredirect(True)
    root.configure(background='#118977')
    root.attributes('-alpha',0.8)
    root.wm_attributes('-topmost','true')
    root.title(PROGRAM_NAME)
    pause_option=IntVar()
    volume_value=StringVar()
    french_word=StringVar()
    english_word=StringVar()
    volume_caption=StringVar()
    pause_caption=StringVar()
    pause_option.set(0)
    volume_value.set('0.7')
    french_word.set(" ")
    english_word.set(" ")
    volume_caption.set("Volume On")
    pause_caption.set("Pause")
    root.resizable(width=False, height=False)
    french=Label(fg='black',textvariable=french_word,relief=GROOVE,width=64,height=3).grid(row=0,padx=1,pady=1,column=0,columnspan=3)
    
    english=Label(textvariable=english_word,relief=GROOVE,width=64,height=6).grid(row=1,padx=1,pady=1,column=0,columnspan=3)
    example_sentence=Label(text="example sentence",relief=GROOVE,width=64,height=3).grid(row=2,padx=1,pady=1,column=0,columnspan=3)
    image_example=Label(text="picture",relief=GROOVE,width=64,height=16).grid(row=3,padx=1,pady=1,column=0,columnspan=6)    
    
    discard_button=Button(text='Discard',command=lambda : user_discard(french_word.get(),english_word.get()),relief=GROOVE,width=20,height=2).grid(row=4,padx=1,pady=1,column=0,sticky='e')
    voice_button=Button(textvariable=volume_caption,command=lambda :threading.Thread(target=volume_on_off).start(),relief=GROOVE,width=20,height=2).grid(row=4,padx=1,pady=1,column=1,sticky='e')
    exit_button=Button(text='Exit',command=lambda :threading.Thread(target=exit_from_application).start(),relief=GROOVE,width=20,height=2).grid(row=4,padx=1,pady=1,column=2,sticky='e')
    pause_button=Button(textvariable=pause_caption,command=pause,relief=GROOVE,width=20,height=2).grid(row=5,padx=1,pady=1,column=0,sticky='e')
    import_data_button=Button(text='Import Data',command=lambda :threading.Thread(target=ask_to_choose_import_file).start(),relief=GROOVE,width=20,height=2).grid(row=5,padx=1,pady=1,column=1,sticky='e')
    settings_button=Button(text='Settings',relief=GROOVE,width=20,height=2).grid(row=5,padx=1,pady=1,column=2,sticky='e')
    placement(root)
    w=threading.Thread(target=get_words).start()
    root.mainloop()
    
    
main()      


    
    
