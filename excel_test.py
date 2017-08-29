from openpyxl import load_workbook

wbook=load_workbook('./wordList.xlsx')
sheetList=wbook.get_sheet_names()
sheet=wbook.get_sheet_by_name(sheetList[0])

wordlist={}
for i in range(1,20):
    if sheet.cell(row=i,column=1).value!=None:
        wordlist['{}'.format(sheet.cell(row=i,column=1).value)]='{}'.format(sheet.cell(row=i,column=2).value)
        
    else:break
    
    
    
for key,value in wordlist.items():
    print(key,value)
    sys.exit()